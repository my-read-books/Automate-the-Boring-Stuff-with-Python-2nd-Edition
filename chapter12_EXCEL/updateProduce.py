import openpyxl

wb = openpyxl.load_workbook("../automate_online-materials/produceSales.xlsx")
sheet = wb["Sheet"]

PRICE_UPDATE = {"Lemon": 3.07,
               "Celery": 1.19,
               "Garlic": 1.27}

for rowNum in range(2, sheet.max_row + 1):
    produceName = sheet.cell(rowNum, 1).value
    if produceName in PRICE_UPDATE:
        sheet.cell(rowNum, 2).value = PRICE_UPDATE[produceName]
wb.save("updatedProduceSales.xlsx")
