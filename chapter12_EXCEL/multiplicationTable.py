from openpyxl import Workbook
from openpyxl.styles import Font
size = int(input("Enter size of table: "))
wb = Workbook()
sheet = wb.active

for row in range(1, size+1):
    sheet.cell(row+1, 1).value = row
    sheet.cell(row+1, 1).font = Font(bold=True)
    for col in range(1, size + 1):
        sheet.cell(1, col+1).value = col
        sheet.cell(1, col+1).font = Font(bold=True)
        sheet.cell(row+1, col+1).value = row * col
wb.save("multiplicationTabel.xlsx")


